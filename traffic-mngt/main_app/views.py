import datetime
import os
import uuid
from django.http import HttpResponse
from django.shortcuts import render

from main_app.forms import AccidentForm
from django.core.files.base import ContentFile
from main_app.models import Accident
from main_app.services import MongoService

from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect
from django.utils.timezone import now
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# Create your views here.
def history_alert_view(request):
    service = MongoService()
    alerts = service.get_latest_alerts(limit_days=5)

    for alert in alerts:
        # Convert detect_at timestamp -> readable
        if "detect_at" in alert:
            try:
                alert["detect_at"] = datetime.datetime.fromtimestamp(alert["detect_at"]).strftime('%Y-%m-%d %H:%M:%S')
            except Exception:
                alert["detect_at"] = str(alert["detect_at"])

    paginator = Paginator(alerts, 10)
    page = request.GET.get('page')
    alerts_page = paginator.get_page(page)
    
    return render(request, 'history_alert.html', {'alerts': alerts_page})


def accident_list_view(request):
    accident_qs = Accident.objects.all().order_by('-created_at')
    paginator = Paginator(accident_qs, 10)  # 10 tai nạn mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'accident/accident_list.html', {
        'accidents': page_obj,
        "now": now()
    })


def accident_detail(request, pk):
    accident = get_object_or_404(Accident, pk=pk)
    return render(request, 'accident/accident_detail.html', {'accident': accident})


def accident_create(request):
    if request.method == 'POST':
        if 'snapshot' in request.FILES:
            original_file = request.FILES['snapshot']
            ext = os.path.splitext(original_file.name)[1] or '.jpg'
            new_filename = f"{request.POST.get('camera_serial', 'camera')}_{uuid.uuid4()}{ext}"

            # Tạo lại file với tên mới
            new_file = ContentFile(original_file.read())
            new_file.name = new_filename
            request.FILES['snapshot'] = new_file

        form = AccidentForm(request.POST, request.FILES)
        if form.is_valid():
            accident = form.save(commit=False)
            accident.confirmed_by = request.user
            accident.save()
            return redirect('accident')
    else:
        form = AccidentForm()
    return render(request, 'accident/accident_form.html', {'form': form, 'is_create': True})


def accident_update(request, pk):
    accident = get_object_or_404(Accident, pk=pk)
    old_snapshot = accident.snapshot if accident.snapshot else None

    if request.method == 'POST':
        if 'snapshot' in request.FILES:
            # Xoá ảnh cũ trên S3 (nếu có)
            if old_snapshot:
                old_snapshot.delete()

            # Đổi tên file mới trước khi gán vào form
            original_file = request.FILES['snapshot']
            ext = os.path.splitext(original_file.name)[1] or '.jpg'
            new_filename = f"{accident.camera_serial}_{uuid.uuid4()}{ext}"
            new_file = ContentFile(original_file.read())
            new_file.name = new_filename
            request.FILES['snapshot'] = new_file

        form = AccidentForm(request.POST, request.FILES, instance=accident)
        if form.is_valid():
            accident = form.save(commit=False)
            accident.confirmed_by = request.user
            accident.save()
            return redirect('accident')
    else:
        form = AccidentForm(instance=accident)

    return render(request, 'accident/accident_form.html', {'form': form})


def accident_delete(request, pk):
    accident = get_object_or_404(Accident, pk=pk)
    if request.method == 'POST':
        accident.delete()
        return redirect('accident')
    return render(request, 'accident/accident_confirm_delete.html', {'accident': accident})

def export_accident_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    filters = Q()
    if start_date:
        filters &= Q(created_at__date__gte=start_date)
    if end_date:
        filters &= Q(created_at__date__lte=end_date)

    if filters:
        queryset = Accident.objects.all().filter(filters).order_by('-created_at')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách tai nạn"

    # Style
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Header row
    headers = ["#", "Thời gian tạo", "Camera Serial", "Người xác nhận"]
    ws.append(headers)

    for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
        for cell in col:
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border

    # Data rows
    for idx, accident in enumerate(queryset, start=1):
        row = [
            idx,
            accident.created_at.strftime('%d/%m/%Y %H:%M:%S'),
            accident.camera_serial,
            accident.confirmed_by.email if accident.confirmed_by else "Chưa xác nhận"
        ]
        ws.append(row)

        for i, value in enumerate(row, start=1):
            cell = ws.cell(row=idx + 1, column=i)
            cell.border = thin_border
            if i == 1:
                cell.alignment = center_align

    # Auto-fit column width
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        adjusted_width = max_length + 2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Response
    filename = f"accidents_{start_date}_to_{end_date}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response